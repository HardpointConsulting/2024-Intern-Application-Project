import streamlit as st
from pypdf import *
import os
from datetime import datetime
st.title("PDF Reader")
st.subheader("Upload your files to extract data!!!")

def save_uploaded_file(uploaded_file, save_path):
    with open(f"docs\\{uploaded_file.name}", "wb") as file:
        file.write(uploaded_file.getbuffer())
    return f"docs\\{uploaded_file.name}"

def entry_db(title,year, journal, author,types, summary):
    import openpyxl

    # Load the existing workbook
    wb = openpyxl.load_workbook('Blank Knowledge Database.xlsx')

    # Select the active sheet
    sheet = wb.active

    # Determine the last row with data in column A
    last_row = sheet.max_row + 1

    # Data to be added

    # Add data to the last row
    sheet.cell(row=last_row, column=1).value = last_row-2
    sheet.cell(row=last_row, column=2).value = title
    sheet.cell(row=last_row, column=3).value = year
    sheet.cell(row=last_row, column=4).value = journal
    sheet.cell(row=last_row, column=5).value = author
    sheet.cell(row=last_row, column=6).value = summary
    sheet.cell(row=last_row, column=7).value = types

    # Save the workbook
    wb.save('Blank Knowledge Database.xlsx')

uploaded_file = st.file_uploader("Upload your file as .pdf",type=["pdf"])






if uploaded_file is not None:
    save_path = os.path.join("docs\\", uploaded_file.name)
    saved_path = save_uploaded_file(uploaded_file, save_path)
    st.success(f"File saved")
    reader = PdfReader(save_path)
    text = ""
    meta = reader.metadata
    number_of_pages = len(reader.pages)
    page = reader.pages[0]
    text = page.extract_text()
    title = pdf_title()
    year = pdf_date()
    journal = pdf_journal()
    author = pdf_author()
    summary = pdf_summary()
    types =""
    for i in author:
        auth=i[:-1]
    timestamp = datetime.fromisoformat(str(year))
    year= timestamp.strftime("%Y-%m-%d")
    author=str(auth)
    if title != "":
        st.text_input("Title",title)
    if year != "":
        st.text_input("Year",year)
    if journal!= "":
        st.text_input("Journal",journal)
    if author != "":
        st.text_input("Auther",author)
    if summary != "":
        st.text_area("Summary",summary)
    if st.button("submit"):
        entry_db(title,year, journal, author,types, summary)
        

