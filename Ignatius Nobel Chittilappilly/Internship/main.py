import streamlit as st
from extractor import *
import re

st.set_page_config(
        page_title="PDF Extractor",
        page_icon="🗃"
)
st.title("PDF Extractor 📰")
st.subheader("Upload your files to extract data!!!")

def entry_db(title,year, journal, author,types, summary):
    import openpyxl

    # Load the existing workbook
    wb = openpyxl.load_workbook('Data Extracted.xlsx')

    # Select the active sheet
    sheet = wb.active

    # Determine the last row with data in column A
    last_row = sheet.max_row + 1

    # Data to be added

    # Add data to the last row
    sheet.cell(row=last_row, column=1).value = last_row-2
    sheet.cell(row=last_row, column=2).value = title
    sheet.cell(row=last_row, column=3).value = year
    sheet.cell(row=last_row, column=4).value = journal
    sheet.cell(row=last_row, column=5).value = author
    sheet.cell(row=last_row, column=6).value = summary
    sheet.cell(row=last_row, column=7).value = types

    # Save the workbook
    wb.save('Data Extracted.xlsx')

def save_uploaded_file(uploaded_file):
    with open(f"docs\\{uploaded_file.name}", "wb") as file:
        file.write(uploaded_file.getbuffer())
    return f"docs\\{uploaded_file.name}"

uploaded_file = st.file_uploader("Upload your file as .pdf",type=["pdf"])
if uploaded_file is not None:
    # save_path = os.path.join("docs\\", uploaded_file.name)
    saved_path = save_uploaded_file(uploaded_file)
    st.success(f"File saved")
    reader = PdfReader(saved_path)
    meta = reader.metadata
    doc_title = st.text_input("Title",meta.title)
    year = meta.creation_date
    doc_year=st.text_input("Year",year.year)
    doc = re.sub('[^A-Za-z0-9() ]+', '', meta.subject)
    doc_journal = st.text_input("Journal",doc)
    type = st.text_input("Type",)
    doc_author = pdf_author(saved_path)
    doc_summary = pdf_summary(saved_path)
    auth = st.text_input("Author",doc_author)
    summary = st.text_area("Document Summary",doc_summary)
    if st.button("Submit"):
        entry_db(doc_title,doc_year, doc_journal, auth,type, summary)
    
    
